from tkinter import font
import PySimpleGUI as sg
from openpyxl import workbook, load_workbook

success = 1

def clt_data():
    wb = load_workbook(filename=locs)
    ws = wb.active
    
    srows = list(ws.iter_rows(swar,smxr,x,ecwas))
    #pprint(srows[0])
    shrow = list(ws.iter_rows(shr,shr,x,ecwas))[0]
    data={}
    for i in srows:
        if i[0].value != None :
            data[i[0].value]={}
            for j in range(scwas-x,len(i)):
                data[i[0].value][shrow[i[j].column-x].value]=(i[j].value)
    return(data)



def dmp_data():
    data = clt_data()
    #pprint(data)
    wbd = load_workbook(filename=locd)
    ws=wbd.active
    dcols = list(ws.iter_cols(xd,xd,srwad,erwad))[0]
    dhrow1 = list(ws.iter_rows(dhr,dhr,scwad,ecwad))[0]
    dhrow ={}
    for u in dhrow1:
        dhrow[u.column] = u
    #pprint(data)
    for i in dcols:
        #print('"'+ i.value +'"')
        if i.value in data:
            row=i.row
            cells=list(ws.iter_rows(row,row,scwad,ecwad))[0]
            for cell in cells:
                if dhrow[cell.column].value in data[i.value]:
                    cell.value=data[i.value][dhrow[cell.column].value]
    wbd.save(filename="C:\\Users\\manis\\Downloads\\Madurai CFA Daily Report - Dump file.xlsx")

sg.theme('Material1')

layout = [[sg.Text('select source:'), sg.Input(key="sloc") ,sg.FileBrowse(), sg.Text('  select dump:'),sg.Input(key="dloc") ,sg.FileBrowse()],
[sg.Text(' ')],
[sg.Text('source data:')],
[sg.Text('heading column in source:'), sg.Input(key='hcs'), sg.Text('heading row in source:'), sg.Input(key='hrs')],
[sg.Text('start row of read aera:'), sg.Input(key='srswa'), sg.Text('max row to read:'),sg.Input(key='mxrw')],
[sg.Text('start col of read aera:'),sg.Input(key='scswa'), sg.Text('max col to read:'),sg.Input(key="mxcl")], 
[sg.Text(' ')],
[sg.Text('dump data:')],
[sg.Text('heading column in dump:'), sg.Input(key='hcd'), sg.Text('heading row in dump:'), sg.Input(key='hrd')],
[sg.Text('start row of dump aera:'), sg.Input(key='srdwa'), sg.Text('max row to read:'),sg.Input(key='mxrwd')],
[sg.Text('start col of dump aera:'),sg.Input(key='scdwa'), sg.Text('max col to dump:'),sg.Input(key="mxcld")], 
[sg.Text(' ')],
[sg.Ok(), sg.Exit(), sg.Text('                                              made by Manish')]]

window = sg.Window('xl data dumper', layout)
while True:
    event,value = window.read()
    if event == sg.WIN_CLOSED or event == 'Exit':
        break
    
    if '' not in (value.values()):
        #sg.popup(value['sloc']+' '+value['dloc']+' '+ value['mxrw'] )
        locs = value['sloc']
        locd = value['dloc']
        
        try:
            #source:
            x = int(value['hcs']) #column with headings in source
            shr = int(value['hrs']) #the row with headings in source
            swar = int(value['srswa']) #start row of working aera in source
            smxr = int(value['mxrw']) #end row of working aera in source
            scwas = int(value['scswa']) #start column of working aera in source
            ecwas = int(value['mxcl']) #end column of working aera in source
            #dump:
            xd = int(value['hcd']) #the column with headings in dump
            dhr = int(value['hrd']) #the row with headings in dump
            srwad = int(value['srdwa']) #start row of working aera in dump
            erwad = int(value['mxrwd']) #end row of working aera in dump
            scwad = int(value['scdwa']) #start column of working aera in dump
            ecwad = int(value['mxcld']) #end column of working aera in dump
        except ValueError:
            sg.popup_error("all inputs must be a numbers (except locations)")
        try:
            dmp_data()
        except:
            success= 0

        if success == 1:
            sg.popup("dumping finished successfully")
        else:
            sg.popup("Error occured while wriing data(check if the file is read only)")
    else:
        sg.popup_error("all fields must be filled")

window.close()